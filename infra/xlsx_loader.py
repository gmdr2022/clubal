from __future__ import annotations

import os
from typing import Callable, List, Optional, Tuple

from openpyxl import load_workbook

from core.models import ClassItem


DEFAULT_SHEET_NAME = "CLUBAL"
EXPECTED_HEADERS = ["DIA", "INICIO", "FIM", "MODALIDADE", "PROFESSOR", "TAG"]

XLSX_STATE_OK = "ok"
XLSX_STATE_MISSING = "missing"
XLSX_STATE_ERROR = "error"


def _normalize_header(v: object) -> str:
    return str(v or "").strip().upper()


def load_classes_from_excel(path: str, sheet_name: str = DEFAULT_SHEET_NAME) -> List[ClassItem]:
    wb = load_workbook(path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Aba '{sheet_name}' não encontrada.")

    ws = wb[sheet_name]

    header_row = None
    headers: List[str] = []

    for r in range(1, 6):
        vals = [_normalize_header(ws.cell(row=r, column=c).value) for c in range(1, 30)]
        if "DIA" in vals and "INICIO" in vals and "FIM" in vals:
            header_row = r
            headers = vals
            break

    if header_row is None:
        header_row = 1
        headers = [_normalize_header(ws.cell(row=1, column=c).value) for c in range(1, 30)]

    def col_idx(name: str) -> Optional[int]:
        name = name.upper()
        if name in headers:
            return headers.index(name) + 1
        return None

    c_dia = col_idx("DIA")
    c_ini = col_idx("INICIO")
    c_fim = col_idx("FIM")
    c_mod = col_idx("MODALIDADE")
    c_prof = col_idx("PROFESSOR")
    c_tag = col_idx("TAG")

    if not (c_dia and c_ini and c_fim and c_mod):
        raise RuntimeError("Cabeçalhos necessários não encontrados na aba CLUBAL.")

    items: List[ClassItem] = []

    for r in range(header_row + 1, ws.max_row + 1):
        dia = str(ws.cell(row=r, column=c_dia).value or "").strip().upper()
        ini = str(ws.cell(row=r, column=c_ini).value or "").strip()
        fim = str(ws.cell(row=r, column=c_fim).value or "").strip()
        modalidade = str(ws.cell(row=r, column=c_mod).value or "").strip()

        if not dia or not ini or not fim or not modalidade:
            continue

        professor = str(ws.cell(row=r, column=c_prof).value or "").strip() if c_prof else ""
        tag = str(ws.cell(row=r, column=c_tag).value or "").strip().upper() if c_tag else ""

        items.append(ClassItem(dia, ini, fim, modalidade, professor, tag))

    return items


def reload_classes_if_needed(
    path: str,
    current_items: List[ClassItem],
    previous_mtime: Optional[float],
    *,
    force: bool = False,
    sheet_name: str = DEFAULT_SHEET_NAME,
    logger: Optional[Callable[[str], None]] = None,
) -> Tuple[List[ClassItem], Optional[float], bool, str]:
    path = str(path or "").strip()

    if not path or not os.path.isfile(path):
        if logger:
            logger(f"[XLSX] Arquivo ausente: {path}")

        changed = bool(current_items) or (previous_mtime is not None)
        return [], None, changed, XLSX_STATE_MISSING

    try:
        mtime = os.path.getmtime(path)

        if not force and previous_mtime is not None and mtime == previous_mtime:
            return current_items, previous_mtime, False, XLSX_STATE_OK

        items = load_classes_from_excel(path, sheet_name=sheet_name)

        if logger:
            logger(f"[XLSX] Excel carregado: {len(items)} itens. mtime={mtime}")

        return items, mtime, True, XLSX_STATE_OK

    except Exception as e:
        if logger:
            logger(f"[XLSX] Falha ao carregar Excel: {type(e).__name__}: {e}")

        return current_items, previous_mtime, False, XLSX_STATE_ERROR